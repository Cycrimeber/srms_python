from tkinter import *
from openpyxl import Workbook
from openpyxl import load_workbook

class Students:

    def __init__(self, master):
        self.master = master
        self.fram = Frame(self.master).pack(fill=BOTH, expand=1)

        btn = Button(self.fram, text="Exit", command=self.exitWindow)
        btn.place(x=100, y=100)

        btnAddStudent = Button(self.fram, text="Save Student Info", command=self.saveStudentInf)
        btnAddStudent.place(x=100, y=50)

        displayButton = Button(self.fram, text='Print student information', command=self.displayStudentInfo)
        displayButton.place(x=100, y=150)
        # self.fram.pack(fill=BOTH, expand=1)

    def exitWindow(self):
        exit()

    def saveStudentInf(self):
        studentDatabase = Workbook()
        sheet = studentDatabase.active
        sheet.title = "Student Scores"

        sheet['A1'] = "Name"
        sheet['B1'] = "CA"
        sheet['C1'] = "Exam"
        sheet['D1'] = "Total"

        sheet['A2'] = "Isa Muhammed"
        sheet['B2'] = "10"
        sheet['C2'] = "30"
        sheet['D2'] = int(sheet['B2'].value) + int(sheet['C2'].value)

        sheet['A3'] = "James Peter"
        sheet['B3'] = "10"
        sheet['C3'] = "30"
        sheet['D3'] = int(sheet['B3'].value) + int(sheet['C3'].value)

        sheet['A4'] = "Ephraim Chukwuebuka Peter"
        sheet['B4'] = "20"
        sheet['C4'] = "30"
        sheet['D4'] = int(sheet['B4'].value) + int(sheet['C4'].value)

        sheet['A5'] = "Emily-King"
        sheet['B5'] = "50"
        sheet['C5'] = "45"
        sheet['D5'] = int(sheet['B5'].value) + int(sheet['C5'].value)
        studentDatabase.save(filename="studentDatabase.xlsx")

        msg = Label(self.fram, text="Save Successfull")
        msg.place(x=100, y=150)
        # self.fram.pack(fill=BOTH, expand=1)

    def displayStudentInfo(self):
        # load workbook
        wb = load_workbook(filename='studentDatabase.xlsx')
        # select wooksheet from workbook
        sheet = wb['Student Scores']
        # loop through workbook and print student information
        student = dict()
        # print(type(sheet.max_row))
        # print(sheet.max_row)
        # print(sheet['A1'].value,sheet['B1'].value,sheet['C1'].value,sheet['D1'].value)
        for i in range(2,sheet.max_row+1):
            var_name = StringVar(self.fram)
            var_name.set(sheet[f'A%d' %int(i)].value)
            var_assessment = StringVar(self.fram)
            var_assessment.set(sheet[f'B%d' % int(i)].value)
            var_exam = StringVar(self.fram)
            var_exam.set(sheet[f'C%d' % int(i)].value)
            var_total = StringVar(self.fram)
            var_total.set(sheet[f'D%d' % int(i)].value)

            l_name = StringVar(self.fram)
            l_name.set(sheet['A1'].value)
            l_assessment = StringVar(self.fram)
            l_assessment.set(sheet['B1'].value)
            l_exam = StringVar(self.fram)
            l_exam.set(sheet['C1'].value)
            l_total = StringVar(self.fram)
            l_total.set(sheet['D1'].value)

            # student['name'] = sheet[f'A%d' %int(i)].value
            # student['assessment'] = sheet[f'B%d' %int(i)].value
            # student['exam'] = sheet[f'C%d' %int(i)].value
            # student['total'] = sheet[f'D%d' %int(i)].value
            # print(student.values())
            # print(sheet)
        # labels
        name = Label(self.fram, textvariable=l_name).place(x=100, y=200)
        assessment = Label(self.fram, textvariable=l_assessment).place(x=200, y=200)
        exam = Label(self.fram, textvariable=l_exam).place(x=230, y=200)
        total = Label(self.fram, textvariable=l_total).place(x=260, y=200)

        # student details
        studentName = Label(self.fram, textvariable=var_name).place(x=100, y=250)
        studentAssessment = Label(self.fram, textvariable=var_assessment).place(x=200, y=250)
        studentExam = Label(self.fram, textvariable=var_exam).place(x=230, y=250)
        studentTotal = Label(self.fram, textvariable=var_total).place(x=260, y=250)

